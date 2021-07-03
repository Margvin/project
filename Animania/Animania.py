#           ИМПОРТ МОДУЛЕЙ

#   Модуль random для поиска псевдослучайного числа
from random import randint

#   Модуль PIL для обработки изображений
from PIL import Image

#   Модуль tkinter для графического интерфейса
from tkinter import ttk, Toplevel, Entry, Text, Label, LabelFrame, Frame, END, CENTER, TOP, X, LEFT, PhotoImage, Tk, \
    BOTH, GROOVE, RIGHT, Button, WORD, SOLID, Checkbutton, IntVar, Canvas, RIDGE, VERTICAL, Y

#   Модуль tkinter.filedialog для запроса файлов
from tkinter.filedialog import askopenfilename

#   Модуль tkinter.messagebox для информационных окон
import tkinter.messagebox as mb

#   Модуль os.path для проверки наличия файлов
import os.path

#   Модуль webbrowser для открытия ссылок в браузере
import webbrowser

#   Модуль openpyxl для работы с excel
import openpyxl

#           ОБЬЯВЛЕНИЕ СПИСКОВ

#   Список стран
spk_stn = ['', 'Абхазия', 'Австралия', 'Австрия', 'Азербайджан', 'Албания', 'Алжир', 'Ангола', 'Андорра',
           'Антигуа и Барбуда', 'Аргентина', 'Армения', 'Афганистан', 'Багамские Острова', 'Бангладеш', 'Барбадос',
           'Бахрейн', 'Белиз', 'Белоруссия', 'Бельгия', 'Бенин', 'Болгария', 'Боливи', 'Босния и Герцеговина',
           'Ботсвана', 'Бразилия', 'Бруней', 'Буркина-Фасо', 'Бурунди', 'Бутан', 'Вануату', 'Ватикан',
           'Великобритания', 'Венгрия', 'Венесуэла', 'Восточный Тимор', 'Вьетнам', 'Габон', 'Гаити', 'Гайана', 'Гамбия',
           'Гана', 'Гватемала', 'Гвинея', 'Гвинея-Бисау', 'Германия', 'Гондурас', 'Гренада', 'Греция', 'Грузия',
           'Дания', 'Джибути', 'Доминика', 'Доминиканская Республика', 'ДР Конго', 'Египет', 'Замбия', 'Зимбабве',
           'Израиль', 'Индия', 'Индонезия', 'Иордания', 'Ирак', 'Иран', 'Ирландия', 'Исландия', 'Испания', 'Италия',
           'Йемен', 'Кабо-Верде', 'Казахстан', 'Камбоджа', 'Камерун', 'Канада', 'Катар', 'Кения', 'Кипр', 'Киргизия',
           'Кирибати', 'Китай', 'КНДР', 'Колумбия', 'Коморские Острова', 'Коста-Рика', "Кот-д'Ивуар", 'Куба', 'Кувейт',
           'Лаос', 'Латвия', 'Лесото', 'Либерия', 'Ливан', 'Ливия', 'Литва', 'Лихтенштейн', 'Люксембург', 'Маврикий',
           'Мавритания', 'Мадагаскар', 'Малави', 'Малайзия', 'Мали', 'Мальдивские Острова', 'Мальта', 'Марокко',
           'Маршалловы Острова', 'Мексика', 'Мозамбик', 'Молдавия', 'Монако', 'Монголия', 'Мьянма', 'Намибия', 'Науру',
           'Непал', 'Нигер', 'Нигерия', 'Нидерланды', 'Никарагуа', 'Новая Зеландия', 'Норвегия', 'ОАЭ', 'Оман',
           'Пакистан', 'Палау', 'Палестина', 'Панама', 'Папуа - Новая Гвинея', 'Парагвай', 'Перу', 'Польша',
           'Португалия', 'Республика Конго', 'Россия', 'Руанда', 'Румыния', 'Сальвадор', 'Самоа', 'Сан - Марино',
           'Сан - Томе и Принсипи', ' Саудовская Аравия', 'Северная Корея', 'Северная Македония',
           'Сейшельские Острова', 'Сенегал', 'Сент - Винсент и Гренадины', 'Сент - Китс и Невис', 'Сент - Люсия',
           'Сербия', 'Сингапур', 'Сирия', 'Словакия', 'Словения', 'Соломоновы Острова', 'Сомали', 'Судан', 'Суринам',
           'США', 'Сьерра - Леоне', 'Таджикистан', 'Таиланд', 'Танзания', 'Того', 'Тонга', 'Тринидад и Тобаго',
           'Тувалу', 'Тунис', 'Туркмения', 'Турция', 'Уганда', 'Узбекистан', 'Украина', 'Уругвай',
           'Федеративные Штаты Микронезии', 'Фиджи', 'Филиппины', 'Финляндия', 'Франция', 'Хорватия', 'ЦАР', 'Чад',
           'Черногория', 'Чехия', 'Чили', 'Швейцария', 'Швеция', 'Шри - Ланка', 'Эквадор', 'Экваториальная Гвинея',
           'Эритрея', 'Эсватини', 'Эстония', 'Эфиопия', 'ЮАР', 'Южная Корея', 'Южная Осетия', 'Южный Судан', 'Ямайка',
           'Япония']

#   Список годов выхода
spk_god = ["", "2030", "2029", "2028", "2027", "2026", "2025", "2024", "2023", "2022", "2021", "2020", "2019",
           "2018", "2017", "2016", "2015", "2014", "2013", "2012", "2011", "2010", "2009", "2008", "2007", "2006",
           "2005", "2004", "2003", "2002", "2001", "2000", "1999", "1998", "1997", "1996", "1995", "1994", "1993",
           "1992", "1991", "1990", "1989", "1988", "1987", "1986", "1985", "1984", "1983", "1982", "1981", "1980",
           "1979", "1978", "1977", "1976", "1975", "1974", "1973", "1972", "1971", "1970", "1969", "1968", "1967",
           "1966", "1965", "1964", "1963", "1962", "1961", "1960", "1959", "1958", "1957", "1956", "1955", "1954",
           "1953", "1952", "1951", "1950", "1949", "1948", "1947", "1946", "1945", "1944", "1943", "1942", "1941",
           "1940", "1939", "1938", "1937", "1936", "1935", "1934", "1933", "1932", "1931", "1930", "1929", "1928",
           "1927", "1926", "1925", "1924", "1923", "1922", "1921", "1920", "1919", "1918", "1917", "1916", "1915",
           "1914", "1913", "1912", "1911", "1910", "1909", "1908", "1907", "1906", "1905", "1904", "1903", "1902",
           "1901", "1900"]

#   Список жанров
spk_jnr = ['', 'Апокалиптика', 'Бара', 'Боевик', 'Боевые искусства', 'Вестерн', 'Гангстерский фильм', 'Детектив',
           'Дзидайгэки', 'Дзёсэй', 'Добуцу', 'Драма', 'Идолы', 'Икудзи', 'Исторический фильм', 'Кайто', 'Киберпанк',
           'Кодомо', 'Комедия', 'Космическая опера', 'Лоликон', 'Махо-сёдзё', 'Мелодрама', 'Меха', 'Мистика', 'Моэ',
           'Музыкальный фильм', 'Научная фантастика', 'Нуар', 'Отаку', 'Парапсихология', 'Паропанк/Стимпанк',
           'Повседневность', 'Политический фильм', 'Полицейский боевик', 'Постапокалиптика', 'Приключения',
           'Психологический триллер', 'Путешествие между мирами', 'Романтика', 'Рэдикоми', 'Самурайский боевик',
           'Семейный', 'Сказка', 'Социальный фильм', 'Спокон', 'Сэйнэн', 'Сэнтай', 'Сёдзё', 'Сёдзё-ай', 'Сёнэн',
           'Сёнэн-ай', 'Сётакон', 'Токусацу', 'Трагедия', 'Трагикомедия', 'Триллер', 'Фантастический фильм',
           'Фильм ужасов', 'Фильм-катастрофа', 'Фэнтези', 'Хентай', 'Этти', 'Юри', 'Яой']

#   Список цифр
spk_cfr = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']


#           ФУНКЦИИ

#   Выбор файла картинки
def fnc_vbr_img(this_window, ent_dob_img):
    filename = askopenfilename(parent=this_window, filetypes=[('JPEG pictures', '*.jpg'), ('PNG pictures', '*.png'),
                                                              ("all files", "*.*")],
                               title='Выберите файл для постера')
    ent_dob_img['bg'] = '#ffffff'
    ent_dob_img.delete(0, END)
    ent_dob_img.insert(0, filename)


#   Закрытие основного окна
def fnc_win_osn_cls(this_window):
    if mb.askokcancel("Выход из программы", "Вы действительно хотите выйти из программы?", parent=this_window):
        this_window.destroy()


#   Закрытие дополнительного окна добавления фильма без сохранения
def fnc_win_dob_cls(this_window):
    if mb.askokcancel("Закрытие окна", "Вы действительно хотите закрыть данное окно?" + '\n\n' +
                                       "Введенные данные не будут сохранены!", parent=this_window):
        this_window.destroy()
        global win_dob
        win_dob = None


#   Закрытие окна информации о фильме
def fnc_win_film_info_cls(this_window):
    this_window.destroy()
    global win_film_info
    win_film_info = None
    global win_all_sps_film
    win_all_sps_film = None

#   Редактирование и сохранения картинки
def fnc_img_obr(ent_dob_img):
    img_pst_1 = Image.open(ent_dob_img.get())
    img_pst_2 = Image.open(ent_dob_img.get())
    new_siz_1 = (400, 500)
    new_siz_2 = (240, 300)
    img_pst_1.thumbnail(new_siz_1)
    img_pst_2.thumbnail(new_siz_2)

    wb = openpyxl.load_workbook('C:/Python/Animania/baza/baza.xlsx')
    sheet = wb.active
    i = 1
    while sheet[f'A{i}'].value is not None:
        i += 1
        nam_csl = i

    wb.save('C:/Python/Animania/baza/baza.xlsx')
    img_pst_1.save(f'images/{nam_csl}.png')
    img_pst_2.save(f'images/{nam_csl}_1.png')
    return f'images/{nam_csl}.png', f'images/{nam_csl}_1.png'


#   Вычисляем рейтинг
def fnc_top_rey():
    wb = openpyxl.load_workbook('C:/Python/Animania/baza/baza.xlsx')
    sheet = wb.active
    i = 1
    while sheet[f'A{i}'].value is not None:
        i += 1
        nam_csl = i

    rey_1 = rey_2 = rey_3 = rey_4 = rey_5 = 0
    rey_1_num = rey_2_num = rey_3_num = rey_4_num = rey_5_num = 0

    for j in range(3, nam_csl):
        if (float(sheet[f'N{j}'].value) + float(sheet[f'P{j}'].value)) / 2 > rey_1:
            rey_5 = rey_4
            rey_4 = rey_3
            rey_3 = rey_2
            rey_2 = rey_1
            rey_5_num = rey_4_num
            rey_4_num = rey_3_num
            rey_3_num = rey_2_num
            rey_2_num = rey_1_num
            rey_1 = (float(sheet[f'N{j}'].value) + float(sheet[f'P{j}'].value)) / 2
            rey_1_num = j

        elif (float(sheet[f'N{j}'].value) + float(sheet[f'P{j}'].value)) / 2 > rey_2:
            rey_5 = rey_4
            rey_4 = rey_3
            rey_3 = rey_2
            rey_5_num = rey_4_num
            rey_4_num = rey_3_num
            rey_3_num = rey_2_num
            rey_2 = (float(sheet[f'N{j}'].value) + float(sheet[f'P{j}'].value)) / 2
            rey_2_num = j

        elif (float(sheet[f'N{j}'].value) + float(sheet[f'P{j}'].value)) / 2 > rey_3:
            rey_5 = rey_4
            rey_4 = rey_3
            rey_5_num = rey_4_num
            rey_4_num = rey_3_num
            rey_3 = (float(sheet[f'N{j}'].value) + float(sheet[f'P{j}'].value)) / 2
            rey_3_num = j

        elif (float(sheet[f'N{j}'].value) + float(sheet[f'P{j}'].value)) / 2 > rey_4:
            rey_5 = rey_4
            rey_5_num = rey_4_num
            rey_4 = (float(sheet[f'N{j}'].value) + float(sheet[f'P{j}'].value)) / 2
            rey_4_num = j

        elif (float(sheet[f'N{j}'].value) + float(sheet[f'P{j}'].value)) / 2 > rey_5:
            rey_5 = (float(sheet[f'N{j}'].value) + float(sheet[f'P{j}'].value)) / 2
            rey_5_num = j

    rey_1_icon = sheet[f'R{rey_1_num}'].value
    rey_2_icon = sheet[f'R{rey_2_num}'].value
    rey_3_icon = sheet[f'R{rey_3_num}'].value
    rey_4_icon = sheet[f'R{rey_4_num}'].value
    rey_5_icon = sheet[f'R{rey_5_num}'].value

    wb.save('C:/Python/Animania/baza/baza.xlsx')
    return rey_1_icon, rey_1_num, rey_2_icon, rey_2_num, rey_3_icon, rey_3_num, rey_4_icon, rey_4_num, rey_5_icon, \
           rey_5_num


#   ОКНО ДЛЯ ПРОСМОТРА ИНФОРМАЦИИ О ФИЛЬМЕ
win_film_info = None


def fnc_win_film(rey_num):
    global win_film_info
    if not win_film_info and not win_dob and not win_all_sps_film:
        #   Создание окна
        win_film_info = Toplevel(win_osn)
        win_film_info.wm_transient(win_osn)
        win_film_info.protocol("WM_DELETE_WINDOW", lambda this_window=win_film_info: fnc_win_film_info_cls(this_window))
        win_film_info.title("Информация о фильме")
        win_film_info.geometry("1000x1000+260+0")
        win_film_info.iconphoto(False, icon_osn)
        win_film_info.config(bg='grey20')
        win_film_info.resizable(False, False)

        #   Открытие файла базы
        wb = openpyxl.load_workbook('C:/Python/Animania/baza/baza.xlsx')
        sheet = wb.active

        #   Создание основного фрейма
        frm_film_info_osn = Frame(win_film_info, bg='grey40', width=990, height=990)
        frm_film_info_osn.place(x=5, y=5)

        #   Создание холст для постера
        cnv_film_info = Canvas(frm_film_info_osn, bg='#ffffff', width=400, height=527)
        cnv_film_info.place(x=3, y=3)

        #   Добавляем постер
        img_film_info = PhotoImage(file=sheet[f'Q{rey_num}'].value)
        image = cnv_film_info.create_image(200, 263, image=img_film_info)

        #   Добавляем полное наименование
        frm_film_info_osn_nam = LabelFrame(frm_film_info_osn, text='Наименование:', bg='#ffffff', fg='black',
                                           relief=RIDGE,
                                           bd=3, font=('Comic Sans MS', '16'), width=576, height=103)
        frm_film_info_osn_nam.place(x=411, y=3)

        txt_film_info_osn_nam = Text(frm_film_info_osn_nam, width=43, height=2, bg='#ffffff', fg='black',
                                     font=('Comic Sans MS', '16'), wrap=WORD)
        txt_film_info_osn_nam.insert(0.0, sheet[f'A{rey_num}'].value)
        txt_film_info_osn_nam.configure(state='disabled')
        txt_film_info_osn_nam.tag_configure("center", justify='center')
        txt_film_info_osn_nam.tag_add("center", 1.0, "end")
        txt_film_info_osn_nam.place(x=3, y=3)

        #   Добавляем дополнительные наименование
        frm_film_info_dop_nam = LabelFrame(frm_film_info_osn, text='Дополнительные наименования:', bg='#ffffff',
                                           fg='black',
                                           relief=RIDGE, bd=3, font=('Comic Sans MS', '16'), width=576, height=170)
        frm_film_info_dop_nam.place(x=411, y=109)

        txt_film_info_dop_nam_1 = Text(frm_film_info_dop_nam, width=43, height=2, bg='#ffffff', fg='black',
                                       font=('Comic Sans MS', '16'), wrap=WORD)
        txt_film_info_dop_nam_1.insert(0.0, sheet[f'B{rey_num}'].value)
        txt_film_info_dop_nam_1.configure(state='disabled')
        txt_film_info_dop_nam_1.tag_configure("center", justify='center')
        txt_film_info_dop_nam_1.tag_add("center", 1.0, "end")
        txt_film_info_dop_nam_1.place(x=3, y=3)

        txt_film_info_dop_nam_2 = Text(frm_film_info_dop_nam, width=43, height=2, bg='#ffffff', fg='black',
                                       font=('Comic Sans MS', '16'), wrap=WORD)
        txt_film_info_dop_nam_2.insert(0.0, sheet[f'C{rey_num}'].value)
        txt_film_info_dop_nam_2.configure(state='disabled')
        txt_film_info_dop_nam_2.tag_configure("center", justify='center')
        txt_film_info_dop_nam_2.tag_add("center", 1.0, "end")
        txt_film_info_dop_nam_2.place(x=3, y=70)

        #   Добавляем произволдство, Страну
        frm_film_info_stn = LabelFrame(frm_film_info_osn, text='Производство:', bg='#ffffff', fg='black',
                                       relief=RIDGE, bd=3, font=('Comic Sans MS', '16'), width=576, height=103)
        frm_film_info_stn.place(x=411, y=280)

        txt_film_info_stn = Text(frm_film_info_stn, width=43, height=2, bg='#ffffff', fg='black',
                                 font=('Comic Sans MS', '16'), wrap=WORD)

        str_film_info_stn = str_film_info_stn_1 = str_film_info_stn_2 = ''
        if sheet[f'D{rey_num}'].value != 'N/A':
            str_film_info_stn_1 = sheet[f'D{rey_num}'].value
        if sheet[f'E{rey_num}'].value != 'N/A':
            str_film_info_stn_2 = sheet[f'E{rey_num}'].value
        if sheet[f'E{rey_num}'].value != 'N/A' and sheet[f'D{rey_num}'].value != 'N/A':
            str_film_info_stn = ', '

        txt_film_info_stn.insert(0.0, str_film_info_stn_1 + str_film_info_stn + str_film_info_stn_2)
        txt_film_info_stn.configure(state='disabled')
        txt_film_info_stn.tag_configure("center", justify='center')
        txt_film_info_stn.tag_add("center", 1.0, "end")
        txt_film_info_stn.place(x=3, y=3)

        #   Добавляем Продолжительнность
        frm_film_info_stn = LabelFrame(frm_film_info_osn, text='Продолжительнность:', bg='#ffffff', fg='black',
                                       relief=RIDGE, bd=3, font=('Comic Sans MS', '16'), width=576, height=73)
        frm_film_info_stn.place(x=411, y=386)

        txt_film_info_stn = Text(frm_film_info_stn, width=43, height=1, bg='#ffffff', fg='black',
                                 font=('Comic Sans MS', '16'), wrap=WORD)
        txt_film_info_stn.insert(0.0, sheet[f'S{rey_num}'].value)
        txt_film_info_stn.configure(state='disabled')
        txt_film_info_stn.tag_configure("center", justify='center')
        txt_film_info_stn.tag_add("center", 1.0, "end")
        txt_film_info_stn.place(x=3, y=3)

        #   Добавляем Год выхода
        frm_film_info_stn = LabelFrame(frm_film_info_osn, text='Год выхода:', bg='#ffffff', fg='black',
                                       relief=RIDGE, bd=1, font=('Comic Sans MS', '16'), width=576, height=73)
        frm_film_info_stn.place(x=411, y=462)

        txt_film_info_stn = Text(frm_film_info_stn, width=43, height=1, bg='#ffffff', fg='black',
                                 font=('Comic Sans MS', '16'), wrap=WORD)
        txt_film_info_stn.insert(0.0, sheet[f'F{rey_num}'].value)
        txt_film_info_stn.configure(state='disabled')
        txt_film_info_stn.tag_configure("center", justify='center')
        txt_film_info_stn.tag_add("center", 1.0, "end")
        txt_film_info_stn.place(x=3, y=3)

        #   Добавляем Жанры
        frm_film_info_jnr = LabelFrame(frm_film_info_osn, text='Жанры:', bg='#ffffff', fg='black',
                                       relief=RIDGE, bd=1, font=('Comic Sans MS', '16'), width=984, height=103)
        frm_film_info_jnr.place(x=3, y=538)

        txt_film_info_jnr = Text(frm_film_info_jnr, width=74, height=2, bg='#ffffff', fg='black',
                                 font=('Comic Sans MS', '16'), wrap=WORD)

        str_film_info_jnr_1 = str_film_info_jnr_2 = str_film_info_jnr_3 = str_film_info_jnr_4 = str_film_info_jnr_5 = \
            str_film_info_jnr_6 = str_film_info_pbl_1 = str_film_info_pbl_2 = str_film_info_pbl_3 = \
            str_film_info_pbl_4 = str_film_info_pbl_5 = ''
        if sheet[f'G{rey_num}'].value != 'N/A':
            str_film_info_jnr_1 = sheet[f'G{rey_num}'].value
            if sheet[f'H{rey_num}'].value != 'N/A':
                str_film_info_jnr_2 = sheet[f'H{rey_num}'].value
                str_film_info_pbl_1 = ', '
                if sheet[f'I{rey_num}'].value != 'N/A':
                    str_film_info_jnr_3 = sheet[f'I{rey_num}'].value
                    str_film_info_pbl_2 = ', '
                    if sheet[f'J{rey_num}'].value != 'N/A':
                        str_film_info_jnr_4 = sheet[f'J{rey_num}'].value
                        str_film_info_pbl_3 = ', '
                        if sheet[f'K{rey_num}'].value != 'N/A':
                            str_film_info_jnr_5 = sheet[f'K{rey_num}'].value
                            str_film_info_pbl_4 = ', '
                            if sheet[f'L{rey_num}'].value != 'N/A':
                                str_film_info_jnr_6 = sheet[f'L{rey_num}'].value
                                str_film_info_pbl_5 = ', '

        txt_film_info_jnr.insert(0.0, str_film_info_jnr_1 + str_film_info_pbl_1 + str_film_info_jnr_2 +
                                 str_film_info_pbl_2 + str_film_info_jnr_3 + str_film_info_pbl_3 + str_film_info_jnr_4 +
                                 str_film_info_pbl_4 + str_film_info_jnr_5 + str_film_info_pbl_5 + str_film_info_jnr_6)

        txt_film_info_jnr.configure(state='disabled')
        txt_film_info_jnr.tag_configure("center", justify='center')
        txt_film_info_jnr.tag_add("center", 1.0, "end")
        txt_film_info_jnr.place(x=7, y=3)

        #   Добавляем описание
        frm_film_info_stn = LabelFrame(frm_film_info_osn, text='Описание:', bg='#ffffff', fg='black',
                                       relief=RIDGE, bd=1, font=('Comic Sans MS', '16'), width=984, height=275)
        frm_film_info_stn.place(x=3, y=645)

        txt_film_info_stn = Text(frm_film_info_stn, width=74, height=8, bg='#ffffff', fg='black',
                                 font=('Comic Sans MS', '16'), wrap=WORD)
        txt_film_info_stn.insert(0.0, sheet[f'V{rey_num}'].value)
        txt_film_info_stn.configure(state='disabled')
        txt_film_info_stn.tag_configure("center", justify='center')
        txt_film_info_stn.tag_add("center", 1.0, "end")
        txt_film_info_stn.place(x=7, y=3)

        #   Добавляем кнопочки
        frm_film_info_stn = Frame(frm_film_info_osn, bg='#ffffff', relief=RIDGE, bd=1, width=984, height=63)
        frm_film_info_stn.place(x=3, y=924)

        #   Ссылка и рейтинг World-art.ru
        frm_film_info_wa = Frame(frm_film_info_stn, bg='#ffffff', relief=RIDGE, bd=1, width=300, height=57)
        frm_film_info_wa.place(x=3, y=3)

        icon_film_info_wa = PhotoImage(file='images/icons/WA_logo.png')
        btn_film_info_wa = Button(frm_film_info_wa, image=icon_film_info_wa, width=181, height=50, bg='#ffffff',
                                  activebackground='#ffffff', bd=0,
                                  command=lambda: webbrowser.open_new_tab(sheet[f'M{rey_num}'].value, ))
        btn_film_info_wa.place(x=3, y=2)

        lbl_film_info_wa_rey = Label(frm_film_info_wa, text=sheet[f'N{rey_num}'].value, width=6, height=1, bd=2,
                                     bg='#ffffff', relief=RIDGE, font=('Comic Sans MS', '20'))
        lbl_film_info_wa_rey.place(x=191, y=6)

        #   Ссылка и рейтинг Кинопоиск
        frm_film_info_kp = Frame(frm_film_info_stn, bg='#ffffff', relief=RIDGE, bd=1, width=169, height=57)
        frm_film_info_kp.place(x=303, y=3)

        icon_film_info_kp = PhotoImage(file='images/icons/KP_logo.png')
        btn_film_info_kp = Button(frm_film_info_kp, image=icon_film_info_kp, width=50, height=50, bg='#ffffff',
                                  activebackground='#ffffff', bd=0,
                                  command=lambda: webbrowser.open_new_tab(sheet[f'O{rey_num}'].value, ))
        btn_film_info_kp.place(x=3, y=2)

        lbl_film_info_kp_rey = Label(frm_film_info_kp, text=sheet[f'P{rey_num}'].value, width=6, height=1, bd=2,
                                     bg='#ffffff', relief=RIDGE, font=('Comic Sans MS', '20'))
        lbl_film_info_kp_rey.place(x=59, y=6)

        #   Возрастное ограничение и нецензурная лексика
        if sheet[f'T{rey_num}'].value == 1:
            frm_film_info_18p = Frame(frm_film_info_stn, bg='#ffffff', relief=RIDGE, bd=1, width=57, height=57)
            frm_film_info_18p.place(x=475, y=3)

            icon_film_info_18p = PhotoImage(file='images/icons/18p_logo.png')
            cnv_film_info_18p = Canvas(frm_film_info_18p, bg='#ffffff', width=45, height=50)
            cnv_film_info_18p.place(x=2, y=1)

            image = cnv_film_info_18p.create_image(26, 25, image=icon_film_info_18p)
            if sheet[f'U{rey_num}'].value == 1:
                frm_film_info_nl = Frame(frm_film_info_stn, bg='#ffffff', relief=RIDGE, bd=1, width=157, height=57)
                frm_film_info_nl.place(x=535, y=3)

                icon_film_info_nl = PhotoImage(file='images/icons/NL_logo.png')
                cnv_film_info_nl = Canvas(frm_film_info_nl, bg='#ffffff', width=150, height=50)
                cnv_film_info_nl.place(x=2, y=2)

                image = cnv_film_info_nl.create_image(75, 25, image=icon_film_info_nl)

        else:
            if sheet[f'U{rey_num}'].value == 1:
                frm_film_info_nl = Frame(frm_film_info_stn, bg='#ffffff', relief=RIDGE, bd=1, width=157, height=57)
                frm_film_info_nl.place(x=475, y=3)

                icon_film_info_nl = PhotoImage(file='images/icons/NL_logo.png')
                cnv_film_info_nl = Canvas(frm_film_info_nl, bg='#ffffff', width=150, height=50)
                cnv_film_info_nl.place(x=2, y=2)

                image = cnv_film_info_nl.create_image(75, 25, image=icon_film_info_nl)

        #   Редактировать фильм
        frm_film_info_rdk = Frame(frm_film_info_stn, bg='#ffffff', relief=RIDGE, bd=0, width=56, height=56)
        frm_film_info_rdk.place(x=920, y=3)

        icon_film_info_rdk = PhotoImage(file='images/icons/rdk_logo.png')
        btn_film_info_rdk = Button(frm_film_info_rdk, image=icon_film_info_rdk, width=50, height=50, bg='#ffffff',
                                   activebackground='#ffffff', bd=0)
        btn_film_info_rdk.place(x=3, y=3)

        wb.save('C:/Python/Animania/baza/baza.xlsx')
        win_film_info.mainloop()


#   ЗАПОЛНЯЕМ МИНИ ИНФОРМАЦИОННОЕ ОКНО
def fnc_prz_win_film(spk_films, win_prz_film):
    for i in spk_films:
        j = int(i)

        #   Открываем файл базы
        wb = openpyxl.load_workbook('C:/Python/Animania/baza/baza.xlsx')
        sheet = wb.active

        #   Создание основного фрейма
        frm_prz_win_film = Frame(win_prz_film, bg='grey40', width=1482, height=310)
        frm_prz_win_film.pack(side=TOP, fill=X, padx=4, pady=4)

        #   Создание дополнительного фрейма
        frm_prz_win_film_1 = Frame(frm_prz_win_film, bg='grey60', width=240, height=300)
        frm_prz_win_film_1.place(x=3, y=3)

        #   Создание холст для постера
        cnv_prz_win_film = Canvas(frm_prz_win_film_1, bg='#ffffff', width=240, height=300)
        cnv_prz_win_film.place(x=2, y=0)

        #   Добавляем постер
        img_prz_win_film = PhotoImage(file=sheet[f'R{j}'].value)
        image = cnv_prz_win_film.create_image(120, 150, image=img_prz_win_film)
        # img_prz_win_film = Image.open(sheet[f'R{i}'].value)
        # image = cnv_prz_win_film.create_image(0, 0, image=img_prz_win_film)

        #   Создание дополнительного фрейма
        frm_prz_win_film_2 = Frame(frm_prz_win_film, bg='grey60', width=730, height=300)
        frm_prz_win_film_2.place(x=246, y=3)

        #   Добавляем полное наименование
        frm_prz_win_film_osn_nam = LabelFrame(frm_prz_win_film_2, text='Наименование:', bg='#ffffff', fg='black',
                                              relief=RIDGE, bd=3, font=('Comic Sans MS', '16'), height=103)
        frm_prz_win_film_osn_nam.pack(side=TOP, fill=X, padx=2, pady=5)

        txt_prz_win_film_osn_nam = Text(frm_prz_win_film_osn_nam, width=91, height=2, bg='#ffffff', fg='black',
                                        font=('Comic Sans MS', '16'), wrap=WORD)
        txt_prz_win_film_osn_nam.insert(0.0, sheet[f'A{j}'].value)
        txt_prz_win_film_osn_nam.configure(state='disabled')
        txt_prz_win_film_osn_nam.tag_configure("center", justify='center')
        txt_prz_win_film_osn_nam.tag_add("center", 1.0, "end")
        txt_prz_win_film_osn_nam.pack(fill=BOTH, padx=2, pady=2)

        #   Добавляем полное наименование
        frm_prz_win_film_ops = LabelFrame(frm_prz_win_film_2, text='Описание:', bg='#ffffff', fg='black',
                                          relief=RIDGE, bd=1, font=('Comic Sans MS', '16'), height=275)
        frm_prz_win_film_ops.pack(side=TOP, fill=X, padx=2, pady=3)

        txt_prz_win_film_ops = Text(frm_prz_win_film_ops, width=91, height=5, bg='#ffffff', fg='black',
                                    font=('Comic Sans MS', '16'), wrap=WORD)
        txt_prz_win_film_ops.insert(0.0, sheet[f'V{j}'].value)
        txt_prz_win_film_ops.configure(state='disabled')
        txt_prz_win_film_ops.tag_configure("center", justify='center')
        txt_prz_win_film_ops.tag_add("center", 1.0, "end")
        txt_prz_win_film_ops.pack(fill=BOTH, padx=2, pady=3)

        #   Создание кнопки
        btn_prz_win_film = Button(cnv_prz_win_film, text='Подробнее...', bg='#ffffff', width=23, height=1, bd=1,
                                  command=lambda: fnc_win_film(j), relief=SOLID, font=('Comic Sans MS', '12'))
        btn_prz_win_film.place(x=0, y=263)

    frm_prz_win_film.mainloop()


#       ОБРАБОТКА ПОЛУЧЕННЫХ ДАННЫХ
#   Получаем переменные
def fnc_dob_prv(win_dob_1, frm_dob_osn, ent_dob_osn_naz, ent_dob_dop_naz_1, ent_dob_dop_naz_2, cmb_dob_stn_1,
                cmb_dob_stn_2, cmb_dob_god, cmb_dob_jnr_1, cmb_dob_jnr_2, cmb_dob_jnr_3, cmb_dob_jnr_4, cmb_dob_jnr_5,
                cmb_dob_jnr_6, ent_dob_sit_1_slk, ent_dob_sit_1_rey, ent_dob_sit_2_slk, ent_dob_sit_2_rey, ent_dob_img,
                ent_dob_prd, var_dob_18p, var_dob_mat, txt_dob_ops):
    #   Очищаем цвет полей
    ent_dob_osn_naz['bg'] = ent_dob_sit_1_slk['bg'] = ent_dob_sit_1_rey['bg'] = ent_dob_dop_naz_1['bg'] = \
        ent_dob_dop_naz_2['bg'] = ent_dob_sit_1_slk['bg'] = ent_dob_sit_1_rey['bg'] = ent_dob_sit_2_slk['bg'] = \
        ent_dob_sit_1_rey['bg'] = ent_dob_img['bg'] = ent_dob_prd['bg'] = '#ffffff'

    #   Проверяем основное наименование
    if ent_dob_osn_naz.get() == '':
        ent_dob_osn_naz['bg'] = 'salmon'
        mb.showwarning("Внимание!", 'Пожалуйста, заполните "Основное название:"', parent=frm_dob_osn)

    #   Проверяем дополнительные наименования
    elif (ent_dob_dop_naz_1.get() == ent_dob_osn_naz.get() or ent_dob_dop_naz_2.get() == ent_dob_osn_naz.get() or
          ent_dob_dop_naz_1.get() == ent_dob_dop_naz_2.get()) and (ent_dob_dop_naz_1.get() != '' or
                                                                   ent_dob_dop_naz_2.get() != ''):
        ent_dob_dop_naz_1['bg'] = 'salmon'
        ent_dob_dop_naz_2['bg'] = 'salmon'
        mb.showwarning("Внимание!", 'В графе "Альтернативные названия:" есть дубли', parent=frm_dob_osn)

    #   Проверяем страны
    elif cmb_dob_stn_1.get() not in spk_stn or cmb_dob_stn_2.get() not in spk_stn:
        mb.showwarning("Внимание!", 'Не коректные данные в графе "Страна:"', parent=frm_dob_osn)

    #   Проверяем год
    elif cmb_dob_god.get() not in spk_god:
        mb.showwarning("Внимание!", 'Не коректные данные в графе "Год:"', parent=frm_dob_osn)

    #   Проверяем жанры
    elif cmb_dob_jnr_1.get() not in spk_jnr or cmb_dob_jnr_2.get() not in spk_jnr or \
            cmb_dob_jnr_3.get() not in spk_jnr or cmb_dob_jnr_4.get() not in spk_jnr or \
            cmb_dob_jnr_5.get() not in spk_jnr or cmb_dob_jnr_6.get() not in spk_jnr:
        mb.showwarning("Внимание!", 'Не коректные данные в графе "Жанр:"', parent=frm_dob_osn)

    #   Проверяем ссылку на World-art.ru
    elif ent_dob_sit_1_slk.get() != '' and ("world-art.ru/animation/animation.php?id=" not in ent_dob_sit_1_slk.get()
                                            and "world-art.ru/cinema/cinema.php?id=" not in ent_dob_sit_1_slk.get()):
        ent_dob_sit_1_slk['bg'] = 'salmon'
        mb.showwarning("Внимание!", 'Пожалуйста, введите ссылку для сайта world-art.ru', parent=frm_dob_osn)

    elif ent_dob_sit_1_slk.get() != '' and ent_dob_sit_1_rey.get() == '':
        ent_dob_sit_1_rey['bg'] = 'salmon'
        mb.showwarning("Внимание!", 'Пожалуйста, введите рейтинг фильма для сайта world-art.ru', parent=frm_dob_osn)

    #   Проверяем рейтинг с World-art.ru
    elif ent_dob_sit_1_rey.get() != '' and len(ent_dob_sit_1_rey.get()) != 6:
        ent_dob_sit_1_rey['bg'] = 'salmon'
        mb.showwarning("Внимание!", 'Пожалуйста, введите рейтинг фильма в формате x.xxxx', parent=frm_dob_osn)

    elif ent_dob_sit_1_rey.get() != '' and ent_dob_sit_1_rey.get()[1] != '.':
        ent_dob_sit_1_rey['bg'] = 'salmon'
        mb.showwarning("Внимание!", 'Пожалуйста, введите рейтинг фильма в формате x.xxxx', parent=frm_dob_osn)

    elif ent_dob_sit_1_rey.get() != '' and (ent_dob_sit_1_rey.get()[0] not in spk_cfr or
                                            ent_dob_sit_1_rey.get()[2] not in spk_cfr or
                                            ent_dob_sit_1_rey.get()[3] not in spk_cfr or
                                            ent_dob_sit_1_rey.get()[4] not in spk_cfr or
                                            ent_dob_sit_1_rey.get()[5] not in spk_cfr):
        ent_dob_sit_1_rey['bg'] = 'salmon'
        mb.showwarning("Внимание!", 'Пожалуйста, введите рейтинг фильма в формате x.xxxx', parent=frm_dob_osn)

    #   Проверяем ссылку на Кинопоиск
    elif ent_dob_sit_2_slk.get() != '' and "kinopoisk.ru/film/" not in ent_dob_sit_2_slk.get():
        ent_dob_sit_2_slk['bg'] = 'salmon'
        mb.showwarning("Внимание!", 'Пожалуйста, введите ссылку для сайта kinopoisk.ru', parent=frm_dob_osn)

    elif ent_dob_sit_2_slk.get() != '' and ent_dob_sit_2_rey.get() == '':
        ent_dob_sit_2_rey['bg'] = 'salmon'
        mb.showwarning("Внимание!", 'Пожалуйста, введите рейтинг фильма для сайта kinopoisk.ru', parent=frm_dob_osn)

    #   Проверяем рейтинг с Кинопоиск
    elif ent_dob_sit_2_rey.get() != '' and len(ent_dob_sit_2_rey.get()) != 5:
        ent_dob_sit_2_rey['bg'] = 'salmon'
        mb.showwarning("Внимание!", 'Пожалуйста, введите рейтинг фильма в формате x.xxx', parent=frm_dob_osn)

    elif ent_dob_sit_2_rey.get() != '' and ent_dob_sit_2_rey.get()[1] != '.':
        ent_dob_sit_2_rey['bg'] = 'salmon'
        mb.showwarning("Внимание!", 'Пожалуйста, введите рейтинг фильма в формате x.xxx', parent=frm_dob_osn)

    elif ent_dob_sit_2_rey.get() != '' and (ent_dob_sit_2_rey.get()[0] not in spk_cfr or
                                            ent_dob_sit_2_rey.get()[2] not in spk_cfr or
                                            ent_dob_sit_2_rey.get()[3] not in spk_cfr or
                                            ent_dob_sit_2_rey.get()[4] not in spk_cfr):
        ent_dob_sit_2_rey['bg'] = 'salmon'
        mb.showwarning("Внимание!", 'Пожалуйста, введите рейтинг фильма в формате x.xxx', parent=frm_dob_osn)

    #   Проверяем наличие картинки
    elif not os.path.exists(ent_dob_img.get()):
        ent_dob_img['bg'] = 'salmon'
        mb.showwarning("Внимание!", 'Пожалуйста, укажите файл для постера', parent=frm_dob_osn)

    #   Проверяем продолжительность фильма
    elif ent_dob_prd.get().count('.') != 0 or not ent_dob_prd.get().isdigit():
        ent_dob_prd['bg'] = 'salmon'
        mb.showwarning("Внимание!", 'Введено некоректное продолжение фильма,\nУкажите целое число!',
                       parent=frm_dob_osn)

    #   Проверяем описание фильма
    elif txt_dob_ops.get(1.0, END) == '' or txt_dob_ops.get(1.0, END) == '\n':
        mb.showwarning("Внимание!", 'Пожалуйста, укажите описание фильма', parent=frm_dob_osn)

    else:
        img_pst, img_icon = fnc_img_obr(ent_dob_img)

        #       СОХРАНЕНИЕ ДАННЫХ
        #   Открываем файл базы
        wb = openpyxl.load_workbook('C:/Python/Animania/baza/baza.xlsx')
        sheet = wb.active
        i = 1
        while sheet[f'A{i}'].value is not None:
            i += 1
            nam_csl = i
        #   Записываем основное наименование
        sheet[f'A{nam_csl}'] = ent_dob_osn_naz.get()

        #   Записываем дополнительные наименования
        if ent_dob_dop_naz_1.get() != '':
            sheet[f'B{nam_csl}'] = ent_dob_dop_naz_1.get()
            if ent_dob_dop_naz_2.get() != '' and ent_dob_dop_naz_2.get() != ent_dob_dop_naz_1.get():
                sheet[f'C{nam_csl}'] = ent_dob_dop_naz_2.get()
            else:
                sheet[f'C{nam_csl}'] = 'N/A'
        elif ent_dob_dop_naz_1.get() == '' and ent_dob_dop_naz_2.get() != '':
            sheet[f'B{nam_csl}'] = ent_dob_dop_naz_2.get()
            sheet[f'C{nam_csl}'] = 'N/A'
        else:
            sheet[f'B{nam_csl}'] = 'N/A'
            sheet[f'C{nam_csl}'] = 'N/A'

        #   Записываем страны
        if cmb_dob_stn_1.get() != '':
            sheet[f'D{nam_csl}'] = cmb_dob_stn_1.get()
            if cmb_dob_stn_2.get() != '' and cmb_dob_stn_2.get() != cmb_dob_stn_1.get():
                sheet[f'E{nam_csl}'] = cmb_dob_stn_2.get()
            else:
                sheet[f'E{nam_csl}'] = 'N/A'
        elif cmb_dob_stn_1.get() == '' and cmb_dob_stn_2.get() != '':
            sheet[f'D{nam_csl}'] = cmb_dob_stn_2.get()
            sheet[f'E{nam_csl}'] = 'N/A'
        else:
            sheet[f'D{nam_csl}'] = 'N/A'
            sheet[f'E{nam_csl}'] = 'N/A'

        #   Записываем год
        if cmb_dob_god.get() != '':
            sheet[f'F{nam_csl}'] = cmb_dob_god.get()
        else:
            sheet[f'F{nam_csl}'] = 'N/A'

        #   Записываем жанры
        if cmb_dob_jnr_1.get() != '':
            sheet[f'G{nam_csl}'] = cmb_dob_jnr_1.get()
            if cmb_dob_jnr_2.get() != '' and cmb_dob_jnr_2.get() != cmb_dob_jnr_1.get():
                sheet[f'H{nam_csl}'] = cmb_dob_jnr_2.get()
                if cmb_dob_jnr_3.get() != '' and cmb_dob_jnr_3.get() != cmb_dob_jnr_1.get() and \
                        cmb_dob_jnr_3.get() != cmb_dob_jnr_2.get():
                    sheet[f'I{nam_csl}'] = cmb_dob_jnr_3.get()
                    if cmb_dob_jnr_4.get() != '' and cmb_dob_jnr_4.get() != cmb_dob_jnr_1.get() and \
                            cmb_dob_jnr_4.get() != cmb_dob_jnr_2.get() and cmb_dob_jnr_4.get() != cmb_dob_jnr_3.get():
                        sheet[f'J{nam_csl}'] = cmb_dob_jnr_4.get()
                        if cmb_dob_jnr_5.get() != '' and cmb_dob_jnr_5.get() != cmb_dob_jnr_1.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_2.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_3.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_4.get():
                            sheet[f'K{nam_csl}'] = cmb_dob_jnr_5.get()
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'L{nam_csl}'] = cmb_dob_jnr_6.get()
                            else:
                                sheet[f'L{nam_csl}'] = 'N/A'
                        else:
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'K{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                    else:
                        if cmb_dob_jnr_5.get() != '' and cmb_dob_jnr_5.get() != cmb_dob_jnr_1.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_2.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_3.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_4.get():
                            sheet[f'J{nam_csl}'] = cmb_dob_jnr_5.get()
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'K{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                        else:
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'J{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                else:
                    if cmb_dob_jnr_4.get() != '' and cmb_dob_jnr_4.get() != cmb_dob_jnr_1.get() and \
                            cmb_dob_jnr_4.get() != cmb_dob_jnr_2.get() and cmb_dob_jnr_4.get() != cmb_dob_jnr_3.get():
                        sheet[f'I{nam_csl}'] = cmb_dob_jnr_4.get()
                        if cmb_dob_jnr_5.get() != '' and cmb_dob_jnr_5.get() != cmb_dob_jnr_1.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_2.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_3.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_4.get():
                            sheet[f'J{nam_csl}'] = cmb_dob_jnr_5.get()
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'K{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                        else:
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'J{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                    else:
                        if cmb_dob_jnr_5.get() != '' and cmb_dob_jnr_5.get() != cmb_dob_jnr_1.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_2.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_3.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_4.get():
                            sheet[f'I{nam_csl}'] = cmb_dob_jnr_5.get()
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'J{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                        else:
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'I{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
            else:
                if cmb_dob_jnr_3.get() != '' and cmb_dob_jnr_3.get() != cmb_dob_jnr_1.get() and \
                        cmb_dob_jnr_3.get() != cmb_dob_jnr_2.get():
                    sheet[f'H{nam_csl}'] = cmb_dob_jnr_3.get()
                    if cmb_dob_jnr_4.get() != '' and cmb_dob_jnr_4.get() != cmb_dob_jnr_1.get() and \
                            cmb_dob_jnr_4.get() != cmb_dob_jnr_2.get() and cmb_dob_jnr_4.get() != cmb_dob_jnr_3.get():
                        sheet[f'I{nam_csl}'] = cmb_dob_jnr_4.get()
                        if cmb_dob_jnr_5.get() != '' and cmb_dob_jnr_5.get() != cmb_dob_jnr_1.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_2.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_3.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_4.get():
                            sheet[f'J{nam_csl}'] = cmb_dob_jnr_5.get()
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'K{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                        else:
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'J{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                    else:
                        if cmb_dob_jnr_5.get() != '' and cmb_dob_jnr_5.get() != cmb_dob_jnr_1.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_2.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_3.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_4.get():
                            sheet[f'I{nam_csl}'] = cmb_dob_jnr_5.get()
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'J{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                        else:
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'I{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                else:
                    if cmb_dob_jnr_4.get() != '' and cmb_dob_jnr_4.get() != cmb_dob_jnr_1.get() and \
                            cmb_dob_jnr_4.get() != cmb_dob_jnr_2.get() and cmb_dob_jnr_4.get() != cmb_dob_jnr_3.get():
                        sheet[f'H{nam_csl}'] = cmb_dob_jnr_4.get()
                        if cmb_dob_jnr_5.get() != '' and cmb_dob_jnr_5.get() != cmb_dob_jnr_1.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_2.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_3.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_4.get():
                            sheet[f'I{nam_csl}'] = cmb_dob_jnr_5.get()
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'J{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                        else:
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'I{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                    else:
                        if cmb_dob_jnr_5.get() != '' and cmb_dob_jnr_5.get() != cmb_dob_jnr_1.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_2.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_3.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_4.get():
                            sheet[f'H{nam_csl}'] = cmb_dob_jnr_5.get()
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'I{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                        else:
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'H{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'H{nam_csl}'] = 'N/A'
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
        else:
            if cmb_dob_jnr_2.get() != '' and cmb_dob_jnr_2.get() != cmb_dob_jnr_1.get():
                sheet[f'G{nam_csl}'] = cmb_dob_jnr_2.get()
                if cmb_dob_jnr_3.get() != '' and cmb_dob_jnr_3.get() != cmb_dob_jnr_1.get() and \
                        cmb_dob_jnr_3.get() != cmb_dob_jnr_2.get():
                    sheet[f'H{nam_csl}'] = cmb_dob_jnr_3.get()
                    if cmb_dob_jnr_4.get() != '' and cmb_dob_jnr_4.get() != cmb_dob_jnr_1.get() and \
                            cmb_dob_jnr_4.get() != cmb_dob_jnr_2.get() and cmb_dob_jnr_4.get() != cmb_dob_jnr_3.get():
                        sheet[f'I{nam_csl}'] = cmb_dob_jnr_4.get()
                        if cmb_dob_jnr_5.get() != '' and cmb_dob_jnr_5.get() != cmb_dob_jnr_1.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_2.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_3.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_4.get():
                            sheet[f'J{nam_csl}'] = cmb_dob_jnr_5.get()
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'K{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                        else:
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'J{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                    else:
                        if cmb_dob_jnr_5.get() != '' and cmb_dob_jnr_5.get() != cmb_dob_jnr_1.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_2.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_3.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_4.get():
                            sheet[f'I{nam_csl}'] = cmb_dob_jnr_5.get()
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'J{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                        else:
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'I{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                else:
                    if cmb_dob_jnr_4.get() != '' and cmb_dob_jnr_4.get() != cmb_dob_jnr_1.get() and \
                            cmb_dob_jnr_4.get() != cmb_dob_jnr_2.get() and cmb_dob_jnr_4.get() != cmb_dob_jnr_3.get():
                        sheet[f'H{nam_csl}'] = cmb_dob_jnr_4.get()
                        if cmb_dob_jnr_5.get() != '' and cmb_dob_jnr_5.get() != cmb_dob_jnr_1.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_2.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_3.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_4.get():
                            sheet[f'I{nam_csl}'] = cmb_dob_jnr_5.get()
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'J{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                        else:
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'I{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                    else:
                        if cmb_dob_jnr_5.get() != '' and cmb_dob_jnr_5.get() != cmb_dob_jnr_1.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_2.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_3.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_4.get():
                            sheet[f'H{nam_csl}'] = cmb_dob_jnr_5.get()
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'I{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                        else:
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'H{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'H{nam_csl}'] = 'N/A'
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
            else:
                if cmb_dob_jnr_3.get() != '' and cmb_dob_jnr_3.get() != cmb_dob_jnr_1.get() and \
                        cmb_dob_jnr_3.get() != cmb_dob_jnr_2.get():
                    sheet[f'G{nam_csl}'] = cmb_dob_jnr_3.get()
                    if cmb_dob_jnr_4.get() != '' and cmb_dob_jnr_4.get() != cmb_dob_jnr_1.get() and \
                            cmb_dob_jnr_4.get() != cmb_dob_jnr_2.get() and cmb_dob_jnr_4.get() != cmb_dob_jnr_3.get():
                        sheet[f'H{nam_csl}'] = cmb_dob_jnr_4.get()
                        if cmb_dob_jnr_5.get() != '' and cmb_dob_jnr_5.get() != cmb_dob_jnr_1.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_2.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_3.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_4.get():
                            sheet[f'I{nam_csl}'] = cmb_dob_jnr_5.get()
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'J{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                        else:
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'I{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                    else:
                        if cmb_dob_jnr_5.get() != '' and cmb_dob_jnr_5.get() != cmb_dob_jnr_1.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_2.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_3.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_4.get():
                            sheet[f'H{nam_csl}'] = cmb_dob_jnr_5.get()
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'I{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                        else:
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'H{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'H{nam_csl}'] = 'N/A'
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                else:
                    if cmb_dob_jnr_4.get() != '' and cmb_dob_jnr_4.get() != cmb_dob_jnr_1.get() and \
                            cmb_dob_jnr_4.get() != cmb_dob_jnr_2.get() and cmb_dob_jnr_4.get() != cmb_dob_jnr_3.get():
                        sheet[f'G{nam_csl}'] = cmb_dob_jnr_4.get()
                        if cmb_dob_jnr_5.get() != '' and cmb_dob_jnr_5.get() != cmb_dob_jnr_1.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_2.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_3.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_4.get():
                            sheet[f'H{nam_csl}'] = cmb_dob_jnr_5.get()
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'I{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                        else:
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'H{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'H{nam_csl}'] = 'N/A'
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                    else:
                        if cmb_dob_jnr_5.get() != '' and cmb_dob_jnr_5.get() != cmb_dob_jnr_1.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_2.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_3.get() and \
                                cmb_dob_jnr_5.get() != cmb_dob_jnr_4.get():
                            sheet[f'G{nam_csl}'] = cmb_dob_jnr_5.get()
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'H{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'H{nam_csl}'] = 'N/A'
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                        else:
                            if cmb_dob_jnr_6.get() != '' and cmb_dob_jnr_6.get() != cmb_dob_jnr_1.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_2.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_3.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_4.get() and \
                                    cmb_dob_jnr_6.get() != cmb_dob_jnr_5.get():
                                sheet[f'G{nam_csl}'] = cmb_dob_jnr_6.get()
                                sheet[f'H{nam_csl}'] = 'N/A'
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'
                            else:
                                sheet[f'G{nam_csl}'] = 'N/A'
                                sheet[f'H{nam_csl}'] = 'N/A'
                                sheet[f'I{nam_csl}'] = 'N/A'
                                sheet[f'J{nam_csl}'] = 'N/A'
                                sheet[f'K{nam_csl}'] = 'N/A'
                                sheet[f'L{nam_csl}'] = 'N/A'

        #   Записываем ссылку на World-art.ru
        if ent_dob_sit_1_slk.get() != '':
            sheet[f'M{nam_csl}'] = ent_dob_sit_1_slk.get()
        else:
            sheet[f'M{nam_csl}'] = 'N/A'

        #   Записываем рейтинг с World-art.ru
        if ent_dob_sit_1_rey.get() != '':
            sheet[f'N{nam_csl}'] = ent_dob_sit_1_rey.get()
        else:
            sheet[f'N{nam_csl}'] = 'N/A'

        #   Записываем ссылку на Кинопоиск
        if ent_dob_sit_2_slk.get() != '':
            sheet[f'O{nam_csl}'] = ent_dob_sit_2_slk.get()
        else:
            sheet[f'O{nam_csl}'] = 'N/A'

        #   Записываем рейтинг с Кинопоиск
        if ent_dob_sit_2_rey.get() != '':
            sheet[f'P{nam_csl}'] = ent_dob_sit_2_rey.get()
        else:
            sheet[f'P{nam_csl}'] = 'N/A'

        #   Записываем расположение картинок
        sheet[f'Q{nam_csl}'] = img_pst
        sheet[f'R{nam_csl}'] = img_icon

        #   Записываем Продолжительность
        sheet[f'S{nam_csl}'] = ent_dob_prd.get()

        #   Записываем наличие контента 18+
        sheet[f'T{nam_csl}'] = str(var_dob_18p.get())

        #   Записываем наличие контента 18+
        sheet[f'U{nam_csl}'] = str(var_dob_mat.get())

        #   Записываем описание фильма
        sheet[f'V{nam_csl}'] = txt_dob_ops.get(1.0, END)

        #   Записываем данные в файл
        wb.save('C:/Python/Animania/baza/baza.xlsx')
        ent_dob_img.delete(0, END)
        win_dob_1.destroy()


#   Добавление нового фильма
win_dob = None


def fnc_dob_flm():
    #   Создание окна
    global win_dob
    if not win_dob and not win_film_info and not win_all_sps_film:
        win_dob = Toplevel(win_osn)
        win_dob.wm_transient(win_osn)
        win_dob.title("Добавление фильма")
        win_dob.geometry("1000x1000+260+0")
        win_dob.protocol("WM_DELETE_WINDOW", lambda this_window=win_dob: fnc_win_dob_cls(this_window))
        win_dob.iconphoto(False, icon_osn)
        win_dob.config(bg='grey40')
        win_dob.resizable(False, False)

        #   Основной фрейм
        frm_dob_osn = Frame(win_dob, bg='grey60', width=990, height=990)
        frm_dob_osn.pack(padx=5, pady=5)

        #   Основное наименование
        frm_dob_osn_naz = LabelFrame(frm_dob_osn, text=' Основное название: ', bg='#ffffff', fg='black', relief=GROOVE,
                                     bd=3, height=50, font=('Comic Sans MS', '16'), width=984)
        frm_dob_osn_naz.pack(side=TOP, fill=X, padx=3, pady=3)
        ent_dob_osn_naz = Entry(frm_dob_osn_naz, bg='#ffffff', fg='black', font=('Comic Sans MS', '16'), justify=CENTER)
        ent_dob_osn_naz.pack(fill=BOTH)

        #   Дополнительные наименования
        frm_dob_dop_naz = LabelFrame(frm_dob_osn, text=' Альтернативные названия: ', bg='#ffffff', fg='black',
                                     relief=GROOVE, bd=5, height=50, font=('Comic Sans MS', '16'), width=984)
        frm_dob_dop_naz.pack(side=TOP, fill=X, padx=3, pady=3)
        ent_dob_dop_naz_1 = Entry(frm_dob_dop_naz, bg='#ffffff', fg='black', font=('Comic Sans MS', '16'),
                                  justify=CENTER)
        ent_dob_dop_naz_1.pack(fill=X)
        ent_dob_dop_naz_2 = Entry(frm_dob_dop_naz, bg='#ffffff', fg='black', font=('Comic Sans MS', '16'),
                                  justify=CENTER)
        ent_dob_dop_naz_2.pack(fill=X)

        #   Страна
        frm_dob_stn_god = Frame(frm_dob_osn, width=984)
        frm_dob_stn_god.pack(side=TOP, fill=X, padx=3, pady=3)
        frm_dob_stn_1 = LabelFrame(frm_dob_stn_god, text='Страна:', bg='#ffffff', fg='black', relief=GROOVE, bd=1,
                                   font=('Comic Sans MS', '16'), width=400)
        frm_dob_stn_1.pack(side=LEFT, padx=1, pady=1)
        cmb_dob_stn_1 = ttk.Combobox(frm_dob_stn_1, values=spk_stn, justify=CENTER, font=('Comic Sans MS', '20'))
        cmb_dob_stn_1.pack()
        frm_dob_stn_2 = LabelFrame(frm_dob_stn_god, text='Страна:', bg='#ffffff', fg='black', relief=GROOVE, bd=1,
                                   font=('Comic Sans MS', '16'), width=400)
        frm_dob_stn_2.pack(side=LEFT, padx=1, pady=1)
        cmb_dob_stn_2 = ttk.Combobox(frm_dob_stn_2, values=spk_stn, justify=CENTER, font=('Comic Sans MS', '20'))
        cmb_dob_stn_2.pack()

        # Год
        frm_dob_god = LabelFrame(frm_dob_stn_god, text='Год:', bg='#ffffff', fg='black', relief=GROOVE, bd=1,
                                 font=('Comic Sans MS', '16'), width=184)
        frm_dob_god.pack(side=RIGHT, padx=1, pady=1)
        cmb_dob_god = ttk.Combobox(frm_dob_god, values=spk_god, justify=CENTER, font=('Comic Sans MS', '20'))
        cmb_dob_god.pack()

        # Жанры
        frm_dob_jnr_all = LabelFrame(frm_dob_osn, text='Жанры:', bg='#ffffff', fg='black', relief=GROOVE, bd=5,
                                     height=50, font=('Comic Sans MS', '16'), width=984)
        frm_dob_jnr_all.pack(side=TOP, fill=X, padx=3, pady=3)
        frm_dob_jnr_1 = Frame(frm_dob_jnr_all)
        frm_dob_jnr_1.pack(fill=X)
        frm_dob_jnr_2 = Frame(frm_dob_jnr_all)
        frm_dob_jnr_2.pack(fill=X)
        cmb_dob_jnr_1 = ttk.Combobox(frm_dob_jnr_1, values=spk_jnr, justify=CENTER, font=('Comic Sans MS', '20'),
                                     width=19)
        cmb_dob_jnr_1.pack(side=LEFT, padx=1, pady=1)
        cmb_dob_jnr_2 = ttk.Combobox(frm_dob_jnr_1, values=spk_jnr, justify=CENTER, font=('Comic Sans MS', '20'),
                                     width=19)
        cmb_dob_jnr_2.pack(side=LEFT, padx=1, pady=1)
        cmb_dob_jnr_3 = ttk.Combobox(frm_dob_jnr_1, values=spk_jnr, justify=CENTER, font=('Comic Sans MS', '20'),
                                     width=19)
        cmb_dob_jnr_3.pack(side=LEFT, padx=1, pady=1)
        cmb_dob_jnr_4 = ttk.Combobox(frm_dob_jnr_2, values=spk_jnr, justify=CENTER, font=('Comic Sans MS', '20'),
                                     width=19)
        cmb_dob_jnr_4.pack(side=LEFT, padx=1, pady=1)
        cmb_dob_jnr_5 = ttk.Combobox(frm_dob_jnr_2, values=spk_jnr, justify=CENTER, font=('Comic Sans MS', '20'),
                                     width=19)
        cmb_dob_jnr_5.pack(side=LEFT, padx=1, pady=1)
        cmb_dob_jnr_6 = ttk.Combobox(frm_dob_jnr_2, values=spk_jnr, justify=CENTER, font=('Comic Sans MS', '20'),
                                     width=19)
        cmb_dob_jnr_6.pack(side=LEFT, padx=1, pady=1)

        #   Сайты
        frm_dob_sit_all = Frame(frm_dob_osn, width=984)
        frm_dob_sit_all.pack(side=TOP, fill=X, padx=3, pady=3)
        frm_dob_sit_1 = Frame(frm_dob_sit_all)
        frm_dob_sit_1.pack(side=TOP, fill=X)
        frm_dob_sit_1_slk = LabelFrame(frm_dob_sit_1, text='Ссылка на world-art.ru:', bg='#ffffff', fg='black',
                                       relief=GROOVE, bd=2, font=('Comic Sans MS', '16'))
        frm_dob_sit_1_slk.pack(side=LEFT)
        ent_dob_sit_1_slk = Entry(frm_dob_sit_1_slk, bg='#ffffff', fg='black', font=('Comic Sans MS', '20'),
                                  justify=CENTER, width=41)
        ent_dob_sit_1_slk.pack(fill=X)
        frm_dob_sit_1_rey = LabelFrame(frm_dob_sit_1, text='Рейтинг:', bg='#ffffff', fg='black', relief=GROOVE, bd=2,
                                       font=('Comic Sans MS', '16'))
        frm_dob_sit_1_rey.pack(side=LEFT)
        ent_dob_sit_1_rey = Entry(frm_dob_sit_1_rey, bg='#ffffff', fg='black', font=('Comic Sans MS', '20'),
                                  justify=CENTER, width=20)
        ent_dob_sit_1_rey.pack(fill=X)

        # 2 Сайт
        frm_dob_sit_2 = Frame(frm_dob_sit_all)
        frm_dob_sit_2.pack(side=TOP, fill=X)
        frm_dob_sit_2_slk = LabelFrame(frm_dob_sit_2, text='Ссылка на kinopoisk.ru:', bg='#ffffff', fg='black',
                                       relief=GROOVE, bd=2, font=('Comic Sans MS', '16'))
        frm_dob_sit_2_slk.pack(side=LEFT)
        ent_dob_sit_2_slk = Entry(frm_dob_sit_2_slk, bg='#ffffff', fg='black', font=('Comic Sans MS', '20'),
                                  justify=CENTER, width=41)
        ent_dob_sit_2_slk.pack(fill=X)
        frm_dob_sit_2_rey = LabelFrame(frm_dob_sit_2, text='Рейтинг:', bg='#ffffff', fg='black', relief=GROOVE, bd=2,
                                       font=('Comic Sans MS', '16'))
        frm_dob_sit_2_rey.pack(side=LEFT)
        ent_dob_sit_2_rey = Entry(frm_dob_sit_2_rey, bg='#ffffff', fg='black', font=('Comic Sans MS', '20'),
                                  justify=CENTER, width=20)
        ent_dob_sit_2_rey.pack(fill=X)

        #   Добавление картинки
        frm_dob_img = LabelFrame(frm_dob_osn, text='Добавьте картинку:', bg='#ffffff', fg='black', relief=GROOVE, bd=2,
                                 font=('Comic Sans MS', '16'), width=984, height=50)
        frm_dob_img.pack(side=TOP, fill=X, padx=3, pady=3)
        ent_dob_img = Entry(frm_dob_img, text='Нажмите "Обзор..." для выбора', bg='#ffffff', fg='grey50',
                            font=('Comic Sans MS', '16'), width=65, justify=CENTER, relief=SOLID, bd=2)
        ent_dob_img.pack(side=LEFT)
        btn_dob_img = Button(frm_dob_img, text='Обзор...', justify=CENTER, bg='#ffffff', font=('Comic Sans MS', '16'),
                             fg='black',
                             width=10, command=lambda this_window=ent_dob_img: fnc_vbr_img(this_window, ent_dob_img))
        btn_dob_img.pack(side=LEFT)

        # Продолжительность, 18+ и мат
        frm_dob_prd_all = Frame(frm_dob_osn, width=984, bg='#ffffff')
        frm_dob_prd_all.pack(side=TOP, fill=X, padx=3, pady=3)

        frm_dob_prd = LabelFrame(frm_dob_prd_all, text='Продолжительность (мин):', bg='#ffffff', fg='black',
                                 relief=GROOVE, bd=2, font=('Comic Sans MS', '16'), width=984, height=50)
        frm_dob_prd.pack(side=LEFT, fill=X, padx=3, pady=3)

        ent_dob_prd = Entry(frm_dob_prd, bg='#ffffff', fg='black', font=('Comic Sans MS', '16'), width=30,
                            justify=CENTER, relief=SOLID, bd=2)
        ent_dob_prd.pack(side=LEFT, fill=X)

        var_dob_18p = IntVar()
        cbn_dob_18p = Checkbutton(frm_dob_prd_all, text='Есть контент 18+', font=('Comic Sans MS', '16'), height=2,
                                  bg='#ffffff', variable=var_dob_18p)
        cbn_dob_18p.pack(side=LEFT, fill=X, padx=30)

        var_dob_mat = IntVar()
        cbn_dob_mat = Checkbutton(frm_dob_prd_all, text='Нецензурная лексика', font=('Comic Sans MS', '16'), height=2,
                                  bg='#ffffff', variable=var_dob_mat)
        cbn_dob_mat.pack(side=LEFT, fill=X, padx=30)

        #   Описание фильма
        frm_dob_ops = LabelFrame(frm_dob_osn, text='Описание фильма', bg='#ffffff', fg='black', relief=GROOVE, bd=2,
                                 font=('Comic Sans MS', '16'), width=984)
        frm_dob_ops.pack(side=TOP, fill=X, padx=2, pady=2)
        txt_dob_ops = Text(frm_dob_ops, bg='#ffffff', fg='black', font=('Comic Sans MS', '16'), height=6, width=2,
                           wrap=WORD)
        txt_dob_ops.pack(fill=X, padx=2, pady=2)

        #   Кнопки обработки
        frm_dob_knp = Frame(frm_dob_osn, bg='grey60', width=984)
        frm_dob_knp.pack(side=TOP, fill=X, padx=3, pady=3)
        btn_dob_sav = Button(frm_dob_knp, text='Записать', justify=CENTER, bg='#ffffff', font=('Comic Sans MS', '16'),
                             fg='black', width=10, command=lambda: fnc_dob_prv(win_dob, frm_dob_osn, ent_dob_osn_naz,
                                                                               ent_dob_dop_naz_1, ent_dob_dop_naz_2,
                                                                               cmb_dob_stn_1, cmb_dob_stn_2,
                                                                               cmb_dob_god, cmb_dob_jnr_1,
                                                                               cmb_dob_jnr_2, cmb_dob_jnr_3,
                                                                               cmb_dob_jnr_4, cmb_dob_jnr_5,
                                                                               cmb_dob_jnr_6, ent_dob_sit_1_slk,
                                                                               ent_dob_sit_1_rey, ent_dob_sit_2_slk,
                                                                               ent_dob_sit_2_rey, ent_dob_img,
                                                                               ent_dob_prd, var_dob_18p, var_dob_mat,
                                                                               txt_dob_ops))
        btn_dob_sav.pack()
        win_dob.mainloop()


#       СПИСОК ФИЛЬМОВ
win_all_sps_film = None


def fnc_all_sps_film():
    global win_all_sps_film
    if not win_all_sps_film and not win_dob and not win_film_info:
        #   Создание окна
        win_all_sps_film = Toplevel(win_osn)
        win_all_sps_film.wm_transient(win_osn)
        win_all_sps_film.protocol("WM_DELETE_WINDOW",
                               lambda this_window=win_all_sps_film: fnc_win_film_info_cls(this_window))
        win_all_sps_film.title("Информация о фильме")
        win_all_sps_film.geometry("1500x1000+10+0")
        win_all_sps_film.iconphoto(False, icon_osn)
        win_all_sps_film.config(bg='grey20')
        win_all_sps_film.resizable(False, False)

        #   Основной фрейм
        frm_all_sps_film_osn = Frame(win_all_sps_film, bg='grey60', width=1490, height=990)
        frm_all_sps_film_osn.pack(padx=5, pady=5, expand=1)

        #   Создаем Холст
        cnv_all_sps_film_osn = Canvas(frm_all_sps_film_osn, width=1460, height=990)
        cnv_all_sps_film_osn.pack(side=LEFT, fill=BOTH, expand=1)

        #   Добавляем скроллбар на Холст
        scrollbar_all_sps_film_osn = ttk.Scrollbar(frm_all_sps_film_osn, orient=VERTICAL,
                                                   command=cnv_all_sps_film_osn.yview)
        scrollbar_all_sps_film_osn.pack(side=RIGHT, fill=Y)

        #   Настраиваем Холст
        cnv_all_sps_film_osn.configure(yscrollcommand=scrollbar_all_sps_film_osn.set)
        cnv_all_sps_film_osn.bind('<Configure>',
                                  lambda e: cnv_all_sps_film_osn.configure(scrollregion=
                                                                           cnv_all_sps_film_osn.bbox("all")))

        #   Создаем дополнительный фрейм
        frm_all_sps_film_osn_1 = Frame(cnv_all_sps_film_osn, width=1465, height=990)

        #   Добавляем дополнительный фрейм на холст0
        cnv_all_sps_film_osn.create_window((0,0), window=frm_all_sps_film_osn_1, anchor="nw")

        #   Открытие файла базы
        wb = openpyxl.load_workbook('C:/Python/Animania/baza/baza.xlsx')
        sheet = wb.active
        i = 1
        while sheet[f'A{i}'].value is not None:
            i += 1
            nam_csl = i

        sps_osn_rnd = []
        for j in range(3, nam_csl):
            sps_osn_rnd += str(j)
        fnc_prz_win_film(sps_osn_rnd, frm_all_sps_film_osn_1)



#       ГРАФИЧЕСКИЙ ИНТЕРФЕЙС


#   Создаем основное окно
win_osn = Tk()
win_osn.title("Animania")
win_osn.geometry("1500x1000+10+0")
icon_osn = PhotoImage(file='images/icons/icon.png')
win_osn.iconphoto(False, icon_osn)
win_osn.protocol("WM_DELETE_WINDOW", lambda this_window=win_osn: fnc_win_osn_cls(this_window))
win_osn.config(bg='grey13')
win_osn.resizable(False, False)

# Создание  верхнего фрейма
frm_osn_top = LabelFrame(win_osn, text=' Лидеры рейтинга: ', bg='grey61', fg='yellow', relief=GROOVE, bd=10,
                         height=300, font=('Comic Sans MS', '14'))
frm_osn_top.pack(side=TOP, fill=X, padx=10, pady=10)

#   Размещаем кнопок на верхнем фрейме
str_osn_top_rey_1_icon, str_osn_top_rey_1_num, str_osn_top_rey_2_icon, str_osn_top_rey_2_num, str_osn_top_rey_3_icon, \
str_osn_top_rey_3_num, str_osn_top_rey_4_icon, str_osn_top_rey_4_num, str_osn_top_rey_5_icon, \
str_osn_top_rey_5_num = fnc_top_rey()

img_rey_1 = PhotoImage(file=str_osn_top_rey_1_icon, height=300)
frm_osn_top_rey_1 = Frame(frm_osn_top)
frm_osn_top_rey_1.pack(side=LEFT, padx=5, pady=5)

btn_osn_top_rey_1 = Button(frm_osn_top_rey_1, image=img_rey_1, justify=CENTER, height=300, width=240,
                           bg='grey61', activebackground='grey61', bd=0,
                           command=lambda: fnc_win_film(str_osn_top_rey_1_num))
btn_osn_top_rey_1.pack()

img_rey_2 = PhotoImage(file=str_osn_top_rey_2_icon, height=300, width=240)
frm_osn_top_rey_2 = Frame(frm_osn_top)
frm_osn_top_rey_2.pack(side=LEFT, padx=5, pady=5)
btn_osn_top_rey_2 = Button(frm_osn_top_rey_2, image=img_rey_2, justify=CENTER, height=300, width=240,
                           bg='grey61', activebackground='grey61', bd=0,
                           command=lambda: fnc_win_film(str_osn_top_rey_2_num))
btn_osn_top_rey_2.pack()

img_rey_3 = PhotoImage(file=str_osn_top_rey_3_icon, height=300, width=240)
frm_osn_top_rey_3 = Frame(frm_osn_top)
frm_osn_top_rey_3.pack(side=LEFT, padx=5, pady=5)
btn_osn_top_rey_3 = Button(frm_osn_top_rey_3, image=img_rey_3, justify=CENTER, height=300, width=240,
                           bg='grey61', activebackground='grey61', bd=0,
                           command=lambda: fnc_win_film(str_osn_top_rey_3_num))
btn_osn_top_rey_3.pack()

img_rey_4 = PhotoImage(file=str_osn_top_rey_4_icon, height=300, width=240)
frm_osn_top_rey_4 = Frame(frm_osn_top)
frm_osn_top_rey_4.pack(side=LEFT, padx=5, pady=5)
btn_osn_top_rey_4 = Button(frm_osn_top_rey_4, image=img_rey_4, justify=CENTER, height=300, width=240,
                           bg='grey61', activebackground='grey61', bd=0,
                           command=lambda: fnc_win_film(str_osn_top_rey_4_num))
btn_osn_top_rey_4.pack()

img_rey_5 = PhotoImage(file=str_osn_top_rey_5_icon, height=300, width=240)
frm_osn_top_rey_5 = Frame(frm_osn_top)
frm_osn_top_rey_5.pack(side=LEFT, padx=5, pady=5)
btn_osn_top_rey_5 = Button(frm_osn_top_rey_5, image=img_rey_5, justify=CENTER, height=300, width=240,
                           bg='grey61', activebackground='grey61', bd=0,
                           command=lambda: fnc_win_film(str_osn_top_rey_5_num))
btn_osn_top_rey_5.pack()

icon_dob = PhotoImage(file='images/icons/dobavit.png')
frm_osn_top_dob = Frame(frm_osn_top, width=500, height=300, bg='white')
frm_osn_top_dob.pack(side=LEFT, padx=5, pady=5)
btn_osn_top_dob = Button(frm_osn_top_dob, image=icon_dob, width=500, height=300, bg='#ffffff',
                         activebackground='#ffffff', command=lambda: fnc_dob_flm())
btn_osn_top_dob.pack(fill=BOTH)

# Создание  фрейма фильтров
frm_osn_fil = LabelFrame(win_osn, text=' Фильтры: ', bg='grey61', fg='yellow', relief=GROOVE, bd=10,
                         height=254, font=('Comic Sans MS', '14'))
frm_osn_fil.pack(side=TOP, fill=X, padx=10)

btn_all_sps_film = Button(frm_osn_fil,text='Все фильмы', justify=CENTER, height=1, width=20, bg='#ffffff',
                          font=('Comic Sans MS', '16'), activebackground='#ffffff', bd=0, command=fnc_all_sps_film)
btn_all_sps_film.place(x=1190, y= 165)

# Создание  фрейма случайного аниме
frm_osn_rnd = LabelFrame(win_osn, text='Случайный фильм: ', bg='grey61', fg='yellow', relief=GROOVE, bd=10,
                         height=310, font=('Comic Sans MS', '14'))
frm_osn_rnd.pack(side=TOP, fill=X, padx=10, pady=10)

#   Открываем файл базы
wb = openpyxl.load_workbook('C:/Python/Animania/baza/baza.xlsx')
sheet = wb.active
i = 1
while sheet[f'A{i}'].value is not None:
    nam_csl = i
    i += 1
sps_osn_rnd = []
sps_osn_rnd += str(randint(3, nam_csl))
fnc_prz_win_film(sps_osn_rnd, frm_osn_rnd)

win_osn.mainloop()
